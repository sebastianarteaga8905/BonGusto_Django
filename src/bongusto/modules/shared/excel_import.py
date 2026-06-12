"""Funciones compartidas para manejar la importacion masiva de datos desde archivos Excel."""

import unicodedata
from zipfile import BadZipFile


def normalizar_encabezado(valor):
    """Limpia y transforma los encabezados del Excel a un formato estandar."""
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")

    for origen, destino in [(" ", "_"), ("-", "_"), ("/", "_")]:
        texto = texto.replace(origen, destino)

    while "__" in texto:
        texto = texto.replace("__", "_")

    return texto.strip("_")


def texto_limpio(valor):
    """Convierte cualquier valor en texto limpio, evitando errores con nulos."""
    if valor is None:
        return ""
    return str(valor).strip()


def obtener_load_workbook():
    """Importa la funcion necesaria de openpyxl para trabajar con archivos Excel."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("Hace falta instalar openpyxl para usar la importacion masiva.")
    return load_workbook


def leer_filas_excel(archivo_excel):
    """Lee un archivo Excel y devuelve sus filas organizadas como lista de diccionarios."""
    if not archivo_excel:
        raise ValueError("Debes seleccionar un archivo Excel para importar.")

    nombre_archivo = str(getattr(archivo_excel, "name", "")).lower()
    if not nombre_archivo.endswith(".xlsx"):
        raise ValueError("Solo se permiten archivos Excel con extension .xlsx.")

    if hasattr(archivo_excel, "seek"):
        archivo_excel.seek(0)

    try:
        libro = obtener_load_workbook()(archivo_excel, data_only=True)
    except BadZipFile as error:
        raise ValueError("El archivo Excel no es valido o esta danado.") from error

    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El archivo Excel esta vacio.")

    encabezados = [normalizar_encabezado(valor) for valor in filas[0]]
    if not any(encabezados):
        raise ValueError("La primera fila del Excel debe contener encabezados.")

    resultado = []
    for fila in filas[1:]:
        if not fila or all(valor in (None, "") for valor in fila):
            continue

        item = {}
        for indice, encabezado in enumerate(encabezados):
            if not encabezado:
                continue
            item[encabezado] = fila[indice] if indice < len(fila) else None

        resultado.append(item)

    if not resultado:
        raise ValueError("El archivo Excel no trae filas de datos para importar.")

    return resultado


__all__ = ["leer_filas_excel", "normalizar_encabezado", "texto_limpio"]